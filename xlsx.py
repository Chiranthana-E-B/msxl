import openpyxl

wb = openpyxl.load_workbook("Book1.xlsx")
#print(type(wb))
#print(wb.sheetnames)

sheets = wb.sheetnames
#print(len(sheets), sheets)

#print(wb.active.title)
#print(wb.sheetnames)

#sh1 = wb["firstpage"]
#print(type(sh1))

#1st way
#print(sh1["B2"].value) or print(wb["firstpage"]["B2"].value)

#2nd way
#print(sh1.cell(1,1).value)

#3rd way
#print(sh1.cell(row=2, column=2).value)

#print(sh1.max_row)
#print(sh1.max_column)

'''
for i in range(1,sh1.max_row+1):
    for j in range(1,sh1.max_column+1):
        print(sh1.cell(i,j).value,end=" ")
    print(" \n")

'''

'''
sh1.cell(row=4,column=1,value="ravi")
sh1.cell(row=4,column=2,value=26)
wb.save("Book1.xlsx")
'''

#rename as sheet name
#wb["firstpage"].title = "firstsheet"
#wb.save("Book1.xlsx")

#PatternFill
from openpyxl.styles import PatternFill

sh1 = wb["firstsheet"]

sh1['A5'].value = "chiru"
sh1["A5"].fill = PatternFill("solid", fgColor = "71FF33")
wb.save("Book1.xlsx")